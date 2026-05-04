"""
CotacaoAutomatica — ReisCorp
Busca preços no Mercado Livre para itens extraídos pelo ExtratoPDF Petronect.
"""

import os, re, threading, time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Constantes ─────────────────────────────────────────────────────────────────
ML_SITE = "MLB"   # Brasil
ML_SEARCH_URL = "https://api.mercadolibre.com/sites/{site}/search"
DOCS_EX = "LI/SECEX · RADAR/RFB · DI-DUIMP/SISCOMEX · NCM+EX Tarifário · Invoice · Packing List · BL/AWB · NF Entrada"
DOCS_Q  = "Inspeção Q Petrobras obrigatória · Plano de Inspeção e Teste (PIT) · Certificado de Qualidade do Fabricante · Relatório de Inspeção no Fornecedor · Certificado de Conformidade com Norma · Laudo de Teste/Ensaio · Norma Aplicável (API/ASME/ISO/NBR) · Coordenar inspetor aprovado pela Petrobras"

# ── Helpers ────────────────────────────────────────────────────────────────────

def extrair_partnumber(descricao: str) -> str:
    """Extrai o part-number após 'Tp:' na descrição SAP."""
    m = re.search(r'Tp:\s*(\S+)', descricao, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def partes_descricao(descricao: str):
    """Divide a descrição SAP por ';' e retorna lista de partes limpas."""
    partes = [p.strip() for p in descricao.split(";") if p.strip()]
    return partes


def tem_ex(descricao: str) -> bool:
    return bool(re.search(r'\bEX\b', descricao, re.IGNORECASE))

def tem_q(descricao: str) -> bool:
    """Detecta exigência de Inspeção Q Petrobras na descrição."""
    return bool(re.search(r'\bIQ\b|\bINSP\.?\s*Q\b|;\s*Q\b|\bQ\s*-\s*INSP|inspeção\s*Q|insp\.?\s*Q', descricao, re.IGNORECASE))


def buscar_ml(query: str, max_results: int = 5):
    """
    Busca no ML e retorna lista de itens filtrados (condition=new, available_quantity>0).
    Cada item: dict com title, price, permalink, available_quantity, attributes.
    """
    if not query or len(query.strip()) < 3:
        return []
    try:
        params = {
            "q": query,
            "condition": "new",
            "limit": 10,
        }
        resp = requests.get(
            ML_SEARCH_URL.format(site=ML_SITE),
            params=params,
            timeout=10,
        )
        if resp.status_code != 200:
            return []
        data = resp.json()
        results = []
        for item in data.get("results", []):
            if item.get("condition") != "new":
                continue
            if item.get("available_quantity", 0) <= 0:
                continue
            results.append(item)
            if len(results) >= max_results:
                break
        return results
    except Exception:
        return []


def melhor_resultado(results, desc_original: str):
    """Retorna o melhor resultado da lista ML (primeiro por ora)."""
    if not results:
        return None
    return results[0]


def classificar_tipo(item_ml: dict, descricao_original: str) -> str:
    """
    Retorna 'Idêntico' se marca ou part-number original aparecem no título ML,
    caso contrário 'Substituto'.
    """
    title = (item_ml.get("title") or "").upper()
    pn = extrair_partnumber(descricao_original).upper()
    # Verifica part-number
    if pn and pn in title:
        return "Idêntico"
    # Verifica partes da descrição (marca = segunda parte geralmente)
    partes = partes_descricao(descricao_original)
    if len(partes) >= 2:
        marca = partes[1].strip().upper()
        if marca and len(marca) > 2 and marca in title:
            return "Idêntico"
    return "Substituto"


def extrair_atributo(item_ml: dict, nome: str) -> str:
    for attr in item_ml.get("attributes", []):
        if attr.get("id", "").upper() == nome.upper() or \
           (attr.get("name") or "").upper() == nome.upper():
            return str(attr.get("value_name") or "")
    return ""


def cotar_item(row: dict, log_callback=None):
    """
    Recebe um dict com: Processo, Item, Descrição, Quantidade, Unidade, Local de Entrega.
    Retorna dict com todos os campos de saída.
    """
    desc = str(row.get("Descrição") or "")
    qty_raw = row.get("Quantidade", 1)
    try:
        quantidade = float(str(qty_raw).replace(",", ".")) if qty_raw else 1
    except Exception:
        quantidade = 1

    flag_ex = "SIM" if tem_ex(desc) else ""
    flag_q = tem_q(row.get('desc', ''))
    docs_importacao = DOCS_EX if flag_ex else ""
    docs_qualidade = DOCS_Q if flag_q else ""

    base = {
        "Processo": row.get("Processo", ""),
        "Item": row.get("Item", ""),
        "Descrição": desc,
        "Quantidade": quantidade,
        "Unidade": row.get("Unidade", ""),
        "Preço Unit. (R$)": "",
        "Preço Total (R$)": "",
        "Fonte": "",
        "Tipo": "",
        "Título Encontrado": "",
        "Link para Compra": "",
        "Estoque Disponível": "",
        "Flag EX": flag_ex,
        "Documentos Importação": docs_importacao,
        "Flag Q (QPF)": flag_q,
        "Documentos Qualidade": docs_qualidade,
    }

    # ── Sequência de buscas ───────────────────────────────────────────────────
    searches = []

    # Busca 1: Part number (após "Tp:")
    pn = extrair_partnumber(desc)
    if pn:
        searches.append(("ML - PN", pn))

    # Busca 2: Marca + produto
    partes = partes_descricao(desc)
    if len(partes) >= 2:
        termo2 = " ".join(partes[:2])
        searches.append(("ML - Marca+Produto", termo2))

    # Busca 3: Produto genérico
    if partes:
        searches.append(("ML - Genérico", partes[0]))

    found = None
    fonte = ""
    for label, query in searches:
        if log_callback:
            log_callback(f"  🔍 {label}: {query[:60]}")
        results = buscar_ml(query)
        if results:
            found = melhor_resultado(results, desc)
            fonte = label
            break
        time.sleep(0.3)  # respeitar rate limit

    if found:
        preco = found.get("price") or 0
        estoque = found.get("available_quantity") or 0
        tipo = classificar_tipo(found, desc)
        titulo = found.get("title", "")
        link = found.get("permalink", "")

        base.update({
            "Preço Unit. (R$)": round(preco, 2),
            "Preço Total (R$)": round(preco * quantidade, 2),
            "Fonte": fonte,
            "Tipo": tipo,
            "Título Encontrado": titulo,
            "Link para Compra": link,
            "Estoque Disponível": estoque,
        })

    return base, (found is not None)


# ── Leitura do Excel de entrada ───────────────────────────────────────────────

def ler_excel(caminho: str) -> list:
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        rows.append(d)
    return rows


# ── Geração do Excel de saída ─────────────────────────────────────────────────

COLUNAS_SAIDA = [
    "Processo", "Item", "Descrição", "Quantidade", "Unidade",
    "Preço Unit. (R$)", "Preço Total (R$)", "Fonte", "Tipo",
    "Título Encontrado", "Link para Compra", "Estoque Disponível",
    "Flag EX", "Documentos Importação",
    "Flag Q (QPF)", "Documentos Qualidade",
]


def gerar_excel(resultados: list, pasta_saida: str) -> str:
    nome = datetime.now().strftime('%Y.%m.%d %H-%M-%S') + ' Cotacao.xlsx'
    caminho = os.path.join(pasta_saida, nome)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotação"

    # Cabeçalho
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="00cc44", size=10)
    for col_idx, col_name in enumerate(COLUNAS_SAIDA, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Dados
    link_font = Font(color="4a90d9", underline="single")
    for row_idx, row in enumerate(resultados, 2):
        for col_idx, col_name in enumerate(COLUNAS_SAIDA, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Link clicável
            if col_name == "Link para Compra" and val:
                cell.hyperlink = val
                cell.font = link_font
            # Zebra
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="2b2b2b")
            cell.alignment = Alignment(wrap_text=True)

    # Larguras
    larguras = {
        "Processo": 14, "Item": 6, "Descrição": 45, "Quantidade": 12,
        "Unidade": 10, "Preço Unit. (R$)": 16, "Preço Total (R$)": 16,
        "Fonte": 18, "Tipo": 12, "Título Encontrado": 45,
        "Link para Compra": 30, "Estoque Disponível": 14,
        "Flag EX": 8, "Documentos Importação": 60, "Flag Q (QPF)": 10, "Documentos Qualidade": 80,
    }
    for col_idx, col_name in enumerate(COLUNAS_SAIDA, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 15)

    wb.save(caminho)
    return caminho


# ── Interface Tkinter ──────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    root.title("Cotação Automática — ReisCorp")
    root.geometry("620x460")
    root.configure(bg="#2b2b2b")
    root.resizable(False, False)

    arquivo_var = tk.StringVar(value="")

    # ── Título
    tk.Label(root, text="🛒 Cotação Automática", bg="#2b2b2b", fg="white",
             font=("Arial", 14, "bold")).pack(pady=(16, 2))
    tk.Label(root, text="Mercado Livre · Busca inteligente por PN, Marca e Produto",
             bg="#2b2b2b", fg="#aaa", font=("Arial", 9)).pack(pady=(0, 12))

    # ── Seleção de arquivo
    frame_sel = tk.Frame(root, bg="#2b2b2b")
    frame_sel.pack(fill="x", padx=24)

    def selecionar():
        p = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if p:
            arquivo_var.set(p)
            label_arquivo.config(text=os.path.basename(p), fg="white")

    btn_sel = tk.Button(frame_sel, text="📂 Selecionar Excel", command=selecionar,
                        bg="#444", fg="white", relief="flat", padx=10, pady=4,
                        font=("Arial", 9))
    btn_sel.pack(side="left")

    label_arquivo = tk.Label(frame_sel, text="Nenhum arquivo selecionado",
                             bg="#2b2b2b", fg="#aaa", font=("Arial", 9))
    label_arquivo.pack(side="left", padx=10)

    # ── Log
    frame_log = tk.Frame(root, bg="#2b2b2b")
    frame_log.pack(fill="both", expand=True, padx=24, pady=(14, 0))

    log_text = tk.Text(frame_log, bg="#1a1a2e", fg="#ccc", font=("Courier", 8),
                       relief="flat", state="disabled", height=14, wrap="word")
    scrollbar = tk.Scrollbar(frame_log, command=log_text.yview)
    log_text.configure(yscrollcommand=scrollbar.set)
    log_text.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    def log(msg: str):
        log_text.config(state="normal")
        log_text.insert("end", msg + "\n")
        log_text.see("end")
        log_text.config(state="disabled")
        root.update_idletasks()

    # ── Progresso
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
    progress_bar.pack(fill="x", padx=24, pady=(8, 0))

    label_status = tk.Label(root, text="Aguardando...", bg="#2b2b2b", fg="#aaa",
                            font=("Arial", 8))
    label_status.pack(pady=(4, 0))

    # ── Botão executar
    def executar():
        arq = arquivo_var.get()
        if not arq:
            messagebox.showwarning("Atenção", "Selecione o arquivo Excel primeiro.")
            return

        btn_exec.config(state="disabled", text="Processando...")
        btn_sel.config(state="disabled")
        root.config(cursor="wait")
        log_text.config(state="normal")
        log_text.delete("1.0", "end")
        log_text.config(state="disabled")
        progress_var.set(0)
        label_status.config(text="Iniciando...")

        def run():
            try:
                log(f"📂 Lendo: {os.path.basename(arq)}")
                rows = ler_excel(arq)
                total = len(rows)
                log(f"📋 {total} itens encontrados\n")

                resultados = []
                cotados = 0
                nao_encontrados = 0

                for i, row in enumerate(rows, 1):
                    desc = str(row.get("Descrição") or "")[:60]
                    log(f"[{i}/{total}] Item {row.get('Item', i)}: {desc}")
                    label_status.config(text=f"Cotando item {i} de {total}...")
                    progress_var.set((i - 1) / total * 100)

                    resultado, encontrado = cotar_item(row, log_callback=log)
                    resultados.append(resultado)
                    if encontrado:
                        cotados += 1
                        log(f"  ✅ R$ {resultado['Preço Unit. (R$)']} — {resultado['Tipo']}")
                    else:
                        nao_encontrados += 1
                        log(f"  ❌ Não encontrado")

                    time.sleep(0.2)  # rate limit gentil

                progress_var.set(100)
                log("\n💾 Gerando Excel de resultado...")
                pasta = os.path.dirname(arq)
                caminho_saida = gerar_excel(resultados, pasta)
                log(f"✅ Salvo: {os.path.basename(caminho_saida)}")

                root.after(0, lambda: concluido(caminho_saida, cotados, nao_encontrados))

            except Exception as e:
                root.after(0, lambda: erro(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def concluido(caminho, cotados, nao_encontrados):
        root.config(cursor="")
        btn_exec.config(state="normal", text="▶ Executar Cotação")
        btn_sel.config(state="normal")
        label_status.config(text="Concluído!", fg="#00cc44")
        messagebox.showinfo(
            "Cotação Concluída",
            f"✅ {cotados} itens cotados\n"
            f"❌ {nao_encontrados} não encontrados\n\n"
            f"📄 Resultado salvo em:\n{caminho}"
        )

    def erro(msg):
        root.config(cursor="")
        btn_exec.config(state="normal", text="▶ Executar Cotação")
        btn_sel.config(state="normal")
        label_status.config(text="Erro!", fg="red")
        messagebox.showerror("Erro", f"❌ {msg}")

    btn_exec = tk.Button(root, text="▶ Executar Cotação", command=executar,
                         bg="#005f99", fg="white", font=("Arial", 11, "bold"),
                         relief="flat", padx=24, pady=8)
    btn_exec.pack(pady=(10, 4))

    tk.Label(root, text="Powered by ReisCorp", fg="#00cc44", bg="#2b2b2b",
             font=("Arial", 8)).pack(side="bottom", pady=6)

    root.mainloop()


if __name__ == "__main__":
    main()
